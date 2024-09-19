"""
입력받은 북마크 리스트를 통해 엑셀 파일을 만듭니다. 북마크의 단계를 기준으로 추출합니다
"""

import os
from natsort import natsorted

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from module.create_log import logging
from module.extract_bookmark import extract_bookmark


def write_excel(wb, input_path, output_path, level):
    """정리된 북마크 리스트를 통해 엑셀로 변환합니다. """
    ws = wb.active
    for root, _, files in os.walk(input_path):
        for file in natsorted(files):
            if file.lower().endswith('.pdf'):
                blank = str(os.path.basename(input_path)).find(' ')
                under_bar = str(os.path.basename(input_path)).find('_')
                if blank != -1 and under_bar != -1:
                    cmt = str(os.path.basename(input_path))[blank+1:under_bar]
                elif blank != -1 and under_bar == -1:
                    cmt = str(os.path.basename(input_path))[blank+1:]
                else:
                    cmt = str(os.path.basename(input_path))
                file_path = os.path.join(root, file)
                last_row = ws.max_row
                tmp = 1
                for item in extract_bookmark(file_path):
                    if len(item) > 1 and item['level'] == level:
                        cnt = last_row + tmp
                        ws.cell(row=cnt, column=2,
                                value=os.path.relpath(
                                    file_path, os.path.dirname(input_path)).split(os.sep)[1])
                        ws.cell(row=cnt, column=4, value=cmt)
                        ws.cell(row=cnt, column=11, value=file)
                        ws.cell(row=cnt, column=6,
                                value=item['parent']['title'])
                        ws.cell(row=cnt, column=9, value=item['title'])
                        tmp += 1

    wb.save(output_path)


def has_header(wb, output_path):
    """엑셀 header가 존재하는지 확인합니다. 존재하지 않을 경우 새로 생성합니다"""
    ws = wb.active
    first_row = ws[1]
    header_exists = any(cell.value for cell in first_row)

    if not header_exists:
        headers = ['일련번호', '기관명', '기관코드', '위원회명', '위원회 코드',
                   '위원(의원)명', '위원(의원) 코드', '질의유형', '질의', '답변 책자 파일명', '파일명']
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
        fill_color = PatternFill(start_color='4f81bd',
                                 end_color='4f81bd', fill_type='solid')
        for col in range(1, 12):
            ws.cell(row=1, column=col).fill = fill_color
    wb.save(output_path)
    return wb


def load_excel(output_path):
    """엑셀을 불러옵니다. 파일이 없는 경우 새로 생성됩니다"""
    if os.path.exists(output_path):
        wb = load_workbook(output_path)
    else:
        try:
            output_excel = output_path
            wb = Workbook()
            wb.save(output_excel)
            wb = load_workbook(output_excel)
        except Exception:  # pylint: disable=W0703
            e = "엑셀 파일 생성 오류"
            logging(e, '', output_path)

    return has_header(wb, output_path)
